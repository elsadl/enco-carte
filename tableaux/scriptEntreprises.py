# -*- coding: utf-8 -*- 

import os
import json
import openpyxl

dict_entreprises = {}

# la fonction pour recuperer les donnees d'un fichier excel
def formatEntreprise(path):
    workbook = openpyxl.load_workbook(path, data_only=True)
    infos = workbook["infos"]
    shareholders = workbook["actionnaires"]
    evolution = workbook["evolution"]

    entreprise = {}

    # on recupere les infos generales
    for row in range(1, infos.max_row + 1):
        intitule = infos.cell(row = row, column = 1).value
        valeur = infos.cell(row = row, column = 2).value
        if valeur is None:
            valeur = ""
        if intitule == "active_in" or intitule == "ressources":
            entreprise[intitule] = valeur.split(", ")
        else:
            entreprise[intitule] = valeur

    # on recupere les shareholders
    entreprise["shareholders"] = []
    for row in range(2, shareholders.max_row + 1):
        i = 0
        shareholder = {}
        shareholder["name"] = shareholders.cell(row = row, column = 1).value
        shareholder["shares"] = shareholders.cell(row = row, column = 2).value
        entreprise["shareholders"].insert(i, shareholder)
        i += 1

    # on recupere les traductions si jamais
    if 'traductions' in workbook.sheetnames:
        translations = workbook["traductions"]
        entreprise["translations"] = {}
        for row in range(1, translations.max_row + 1):
            intitule = translations.cell(row = row, column = 1).value
            valeur = translations.cell(row = row, column = 2).value
            entreprise["translations"][intitule] = valeur

    # on recupere les chiffres par annee
    entreprise["evolution"] = {}

    # en bouclant d'abord sur les colonnes (annees)
    for col in range(2, evolution.max_column + 1):
        year = evolution.cell(row = 1, column = col).value
        entreprise["evolution"][year] = {}
        entreprise["evolution"][year]["year"] = year
        # puis sur les lignes (categories)
        for row in range(2, evolution.max_row + 1):
            category = evolution.cell(row = row, column = 1).value
            entreprise["evolution"][year][category] = evolution.cell(row = row, column = col).value
        
    # print(entreprise)
    return entreprise

# on boucle sur le repertoire entreprises

# ici le chemin du repertoire dans lequel boucler
directory = "data/entreprises"
entreprises = []
for filename in os.listdir(directory):
    # et pour chaque fichier excel on lance la fonction de recuperation des donnees
    if filename.endswith(".xlsx"):
        # on met tout ca dans un objet
        entreprise = formatEntreprise(os.path.join(directory, filename))
        entreprises.append(entreprise)
        dict_entreprises[entreprise["id"]] = entreprise["nom"]

    else:
        continue

# et on exporte au format json
with open('data/entreprises.json', 'w') as outfile:
    json.dump(entreprises, outfile)

with open('data/dict_entreprises.json', 'w') as outfile:
    json.dump(dict_entreprises, outfile)

