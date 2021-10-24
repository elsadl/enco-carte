import json
import openpyxl

path = "data/dataPays.xlsx"

workbook = openpyxl.load_workbook(path)
sheet = workbook.worksheets[0]

pays = {}
traductions = []
country_names = {}

for col in range(2, sheet.max_column + 1):
    name = sheet.cell(row = 1, column = col).value
    iso = sheet.cell(row = 2, column = col).value
    if iso not in pays:
        pays[iso] = {}
        country_names[name] = {}
    for row in range(1, sheet.max_row + 1):
        intitule = sheet.cell(row = row, column = 1).value
        valeur = sheet.cell(row = row, column = col).value
        if intitule == "iso" or intitule == "country":
            pays[iso][intitule] = valeur
        if intitule.startswith("name_"):
            if intitule not in traductions:
                traductions.append(intitule)
            pays[iso][intitule] = valeur
            country_names[name][intitule] = valeur
        elif intitule == "type":
            pays[iso][valeur] = {}
            for inside_row in range(4, sheet.max_row + 1):
                inside_intitule = sheet.cell(row = inside_row, column = 1).value
                inside_valeur = sheet.cell(row = inside_row, column = col).value
                if inside_intitule == "companies_id":
                    if inside_valeur is None:
                        pays[iso][valeur]["companies"] = []
                    else:
                        pays[iso][valeur]["companies"] = inside_valeur.split(", ")
                elif inside_intitule == "companies":
                    continue
                else:
                    pays[iso][valeur][inside_intitule] = inside_valeur
        else:
            continue

with open("data/nuts.json") as jsonFile:
    nuts = json.load(jsonFile)
    jsonFile.close()

nutsEU = nuts["objects"]["nutsrg"]
nutsOthers = nuts["objects"]["cntrg"]

for country in nutsEU["geometries"]:
    country["properties"]["data"] = False

for country in nutsEU["geometries"]:
    for el in pays:
        if pays[el]["iso"] == country["properties"]["id"]:
            country["properties"]["data"] = True
            country["properties"]["name"] = pays[el]["country"]
            country["properties"]["hospitals"] = pays[el]["Hospitals"]
            country["properties"]["nursing_homes"] = pays[el]["Nursing Homes"]
            for lang in traductions:
                country["properties"][lang] = pays[el][lang]

with open('data/nuts_with_data.json', 'w') as outfile:
    json.dump(nuts, outfile)

with open('data/dict_pays.json', 'w') as outfile:
    json.dump(country_names, outfile)
